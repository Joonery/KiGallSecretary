# @ Joonery
# 개념글 뉴스레터 프로그램


### BASIC LOGIC
# 1) 해당 갤러리의 코드를 알아내서, 접속한다.
# 2) 받아온 html 코드를 분석하여, 해당 날짜에 올라온 개념글의 정보를 parcing. (제목, 시간, 추천)
# 3-1) 모든 pacing된 정보를 f.open()으로 txt파일 또는 excel 파일에 정리하여 export.


### ISSUE
# issue #9 : max가 무한에 가까운 숫자고, 페이지를 더 가지 못하는 경우에는 반복문을 탈출하도록 수정할 것.
# issue #8 : 보이스가 있는 경우에는 어떻게 추출해야 하는가? gall_tit ub-word voice_tit 이고, 이건 .text가 안 뽑힌다.
# issue #-2 : days를 활용해서 날짜별 스크래핑도 가능하게 만들기.
# issue #0 : 마이너 갤러리 말고 일반 갤에서도 작동하도록 대응.
# issue #5 : 누구에 관련된 글인지 유형을 알려준다.
# issue #3 : GUI로 동작
# issue #4 : Build to an exe file.
# issue #7 : 제목 파싱이 불안정함. 제목을 따면 뒤에 [36] 댓글도 같이 달려오는데, 이걸 str로 바꿔서 [ 전까지만 따는 방식으로 바꾸기.


### DONE
# issue #-3 : keyword가 있는 것만 추출하기.
# issue #2 : 내용을 메일로 전송
# issue #1 : 내용을 html을 이용해 예쁘게 정리
# issue #-1 : 시간을 더욱 정확하게 표시


from bs4 import BeautifulSoup as bs             # html 따오기
import requests as rq                           # 웹 리퀘스트
import time                                     # 시간
from datetime import datetime, timedelta        # 시간
import os                                       # desktop 경로 알아내기
from openpyxl import Workbook                   # 엑셀
import smtplib                                  # 메일 smtp서버 연동
from email.mime.text import MIMEText            # 메일 규격
from email.mime.multipart import MIMEMultipart  # html 첨부

import html_template as ht                      # html 내역


class KgallAlimi() :

    # 생성자
    def __init__(self, gcode, list_keywords, maxpost=100, export_type ="" , path="", filename="Vtubernews_", mail_addr="") :
        
        # parameters ======================================================================================
        self.gcode = gcode                      # 갤러리 코드
        self.list_keywords = list_keywords      # 검색할 키워드
        self.maxpost = maxpost                  # 최대 수집할 글의 개수
        self.export_type = export_type          # 내보낼 파일 타입 (txt, print, xlsx, email)
        self.path = path                        # 내보낼 경로 (default : desktop)
        # 내보낼 이름 (default : Vtubernews_날짜_시간)
        self.filename = filename + str(datetime.today().strftime('%m_%d_%H_%M'))
        self.mail_info = {'sender' : mail_addr,
                          'password' : "",
                          'rcver' : mail_addr,
                          'subject' : self.filename,
                        }

        # self.days = days                        # 며칠 전까지 수집할 것인가 (현재는 사용하지 않음)

        # parameters ======================================================================================
        self.pagenum = 0                                                        # 개념글 페이지 iteration
        self.today = datetime.today()                                           # 오늘 날짜
        self.user_agent = {'User-agent': 'Mozilla/5.0'}                         # 접속설정
        self.urlbase = "https://gall.dcinside.com/mgallery/board/lists/?id="    # 접속할 개념글 주소 베이스
        self.urlbase_view = "https://gall.dcinside.com/m/"                      # 내보낼 개념글 주소 베이스

        # performance test =================================================================================
        self.time_start = None

    ### 다음 개념글 페이지의 주소를 리턴한다.
    def get_next_page(self) :
        # 조합 방식 : https://gall.dcinside.com/mgallery/board/lists/?id= + kizunaai + &page= + 1 + &exception_mode=recommend
        self.pagenum += 1
        next_page_url = self.urlbase + self.gcode + "&page=" + str(self.pagenum) + "&exception_mode=recommend"
        return next_page_url

    ### URL에서 html을 받아온다.
    def get_html(self, url):
        _html = ""
        suc = False
        
        # 성공할 때까지 반복하며 html을 가져옴.
        while(suc == False):
            
            try:
                response = rq.get(url,headers=self.user_agent)
            
            except rq.exceptions.RequestException as e:
                time.sleep(3)
                continue

            if response.status_code == 200:
                suc = True
                _html = response.text

            else:
                suc = True
                _html = "<tbody><td>잘못된 주소 입니다.</td></tbody>"

        return _html

    ### 오늘의 날짜를 04.09 의 string으로 return.
    def get_today(self) :
        return str(self.today.strftime('%m.%d'))

    ### n일 전의 날짜를 04.09 의 string 형태로 return.
    def get_beforeday(self, nday) :
        return str((datetime.today()-timedelta(days=nday)).strftime('%m.%d'))

    ### html에서 얻어온 데이터를 파싱해 집어넣는다.
    def parse_data(self, html, newposts) :
        
        # 해당 페이지의 정보를 html로 읽어서
        soup = bs(html, 'html.parser')

        # 해당 페이지에 있는 모든 글의 정보를 리스트에 저장
        rawposts = soup.find("tbody").find_all("tr", class_="ub-content us-post")

        # 모든 글들의 리스트에서 필요한 정보만 추출.
        for rawpost in rawposts :

            # 만일 공지나 이벤트인 경우는 넘김.
            if (rawpost.find("td", class_="gall_subject").text == "이벤트") or (rawpost.find("td", class_="gall_subject").text == "공지") or (rawpost.find("td", class_="gall_subject").text == "설문") :
                continue

            # 보이스가 있는 경우는 넘김
            if rawpost.find("td", class_="gall_tit ub-word voice_tit") :
                continue

            # 해당한 키워드가 있는 경우에만
            if self.has_keyword(rawpost) :

                # newpost = [4.9 , 6809, 74, 냠냠, ㅇㅇ, 링크]
                newpost = []

                newpost.append(rawpost.find("td", class_="gall_date").text) # 날짜
                newpost.append(rawpost.find("td", class_="gall_count").text) # 조회수
                newpost.append(rawpost.find("td", class_="gall_recommend").text) # 추천수
                newpost.append(rawpost.find("td", class_="gall_tit ub-word").text[:-5][1:]) # 제목
                newpost.append(self.urlbase_view + self.gcode + "/" + rawpost.find("td", class_="gall_num").text) # 링크 

                # 글 모음에 집어넣음.
                newposts.append(newpost)

            # 키워드가 없는 경우는 넘김
            else :
                continue

            # 만일 최대 개수까지 채웠으면 함수 종료
            if self.is_full(newposts) :
                return

    ### 글의 제목에 지정된 키워드가 있는지를 판정한다.
    def has_keyword(self, title) :
        for word in self.list_keywords :
            if word in title.find("td", class_="gall_tit ub-word").text[:-5][1:] :
                return True
        return False

    ### 파싱한 정보가 최대 개수를 넘었는지 판정한다.
    def is_full(self, newposts) :
        return (len(newposts) >= int(self.maxpost))

    ### 저장할 파일 경로를 반환한다. (디폴트 바탕화면)
    def get_filepath(self, extension) :
        
        # 넘어온 값이 있으면
        if self.path : 
            return self.path + self.filename + extension
        
        # 넘어온 값이 없으면
        else : 
            filename = "\\" + self.filename + extension
            filepath = os.path.join(os.path.expanduser('~'),'Desktop') + filename 
            return filepath

    ### 파싱한 내용을 txt file로 export. 
    def export_txt(self, newposts) :
        
        with open(self.get_filepath(".txt"), "w") as f:
            f.write(newposts) # 링크    
            f.close()

    ### 파싱한 내용을 xlsx file로 export. 
    def export_xlsx(self, newposts) :

        write_wb = Workbook()
        write_ws = write_wb.create_sheet('시트1')
        write_ws = write_wb.active

        # 첫줄은 스키마
        write_ws.cell(row=1, column=1).value = '날짜'
        write_ws.cell(row=1, column=2).value = '조회수'
        write_ws.cell(row=1, column=3).value = '추천수'
        write_ws.cell(row=1, column=4).value = '제목'
        write_ws.cell(row=1, column=5).value = '링크'


        # 내용 입력    
        for i in range(len(newposts)) :
            for j in range(len(newposts[i])) :
                write_ws.cell(row=i+2, column=j+1).value = newposts[i][j]
                if j==4 :
                    write_ws.cell(row=i+2, column=j+1).style = "Hyperlink"

        # 너비조정
        # write_ws.column_dimensions['제목'].width = 55
        # write_ws.column_dimensions['링크'].width = 55

        # 저장
        write_wb.save(self.get_filepath(".xlsx"))

    # 파싱한 내용을 mail로 export.
    def export_mail(self, newposts) :
        
        # pw를 채워넣음
        self.mail_info['password'] = input('PW : ')

        # 메일 구성 (MIMEText)
        msg = MIMEMultipart("alternative")
        msg.set_charset('utf-8')

        # 정보 불러오기
        msg['Subject'] = self.mail_info['subject']
        msg['From'] = self.mail_info['sender']
        msg['To'] = self.mail_info['rcver']

        # html쓰기
        bodyPart = MIMEText(self.trans_html(newposts), 'html', 'utf-8')
        msg.attach( bodyPart )

        # 메일 서버 연결
        s=smtplib.SMTP( "smtp.gmail.com" , 587 )
        s.starttls() #TLS
        s.login( self.mail_info['sender'] , self.mail_info['password'] )
        s.sendmail( self.mail_info['sender'], self.mail_info['rcver'], msg.as_string() )
        s.close()

        print('Successfully Sent!')

    # 파싱한 내용을 html로 변환한다.
    def trans_html(self, newposts) :

        # html_tempalte.py의 내용을 불러온 후
        head = ht.head
        tail = ht.tail
        mid1 = ht.mid1
        mid2 = ht.mid2

        # 소식을 결합하여
        for post in newposts :
            if newposts.index(post)%2 == 0 :
                head += mid1.format(date=post[0], title=post[3], rec=post[2], click=post[1], link=post[4])
            else :
                head += mid2.format(date=post[0], title=post[3], rec=post[2], click=post[1], link=post[4])

        # 완성된 html로 리턴
        body = head + tail
        return body

    ### 파싱한 리스트를 콘솔창에 프린트(확인용)
    def print_list(self, newposts) :
        for post in newposts :
            print(post[0])
            print("조회수 : {} / 추천수 : {}".format(post[1], post[2]))
            print(post[3])
            print(post[4])
            print()

    ### 작업 처리시간을 반환한다.
    def get_perf_time(self) :
        return time.time() - self.time_start

    # 메인 함수.
    def start(self) :       

        # 0) 초기화
        newposts = []
        self.time_start = time.time()

        # 1) 글을 모으기
        while not self.is_full(newposts) :                      # 정해진 개수만큼 수집하기 전까지는
            newurl = self.get_next_page()                       # 다음 페이지의 주소를 구해
            self.parse_data(self.get_html(newurl), newposts)    # 받아온 정보를 newposts에 파싱
        
        # 2) 파일로 내보내기
        if self.export_type == 'xlsx' :     # 엑셀로
            self.export_xlsx(newposts)
        elif self.export_type == 'print' :  # 콘솔창으로
            self.print_list(newposts)
        elif self.export_type == 'mail' :   # 메일로
            self.export_mail(newposts)
        else :
            print('알 수 없는 타입입니다.')

        # 3) 퍼포먼스 측정
        print("소요 시간 : {} sec".format(round(self.get_perf_time(), 3)))