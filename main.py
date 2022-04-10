# @ Joonery
# 개념글 뉴스레터 프로그램



# 기본 로직

# 1) 해당 갤러리의 코드를 알아내서, 접속한다.
# 2) 받아온 html 코드를 분석하여, 해당 날짜에 올라온 개념글의 정보를 parcing.
                                                # (제목, 시간, 추천)
# 3-1) 모든 pacing된 정보를 f.open()으로 txt파일 또는 excel 파일에 정리하여 export.

# issue #8 : 보이스가 있는 경우에는 어떻게 추출해야 하는가? gall_tit ub-word voice_tit 이고, 이건 .text가 안 뽑힌다.
# issue #-2 : days를 활용해서 날짜별 스크래핑도 가능하게 만들기.
# issue #0 : 마이너 갤러리 말고 일반 갤에서도 작동하도록 대응.
# issue #5 : 누구에 관련된 글인지 유형을 알려준다.
# issue #4 : Build to an exe file.
# issue #3 : GUI로 동작
# issue #7 : 제목 파싱이 불안정함. 제목을 따면 뒤에 [36] 댓글도 같이 달려오는데, 이걸 str로 바꿔서 [ 전까지만 따는 방식으로 바꾸기.


# 해낸 것들 모음
# issue #-3 : keyword가 있는 것만 추출하기.
# issue #2 : 내용을 메일로 전송
# issue #1 : 내용을 html을 이용해 예쁘게 정리
# issue #-1 : 시간을 더욱 정확하게 표시


from bs4 import BeautifulSoup as bs             # html 따오기
import requests as rq                           # 웹 리퀘스트
import time                                     # 시간
from datetime import datetime, timedelta        # 시간
import os                                       # desktop 경로 알아내기
from openpyxl import Workbook                   # 엑셀
import smtplib                                  # 메일 smtp서버 연동
from email.mime.text import MIMEText            # 메일 규격
from email.mime.multipart import MIMEMultipart  # html첨부


class KgallAlimi() :

    # 생성자
    def __init__(self, gcode, list_keywords, maxpost=100, export_type ="" , path="", filename="Vtubernews_", mail_addr="") :
        
        # parameters ======================================================================================
        self.gcode = gcode                      # 갤러리 코드
        self.list_keywords = list_keywords      # 검색할 키워드
        self.maxpost = maxpost                  # 최대 수집할 글의 개수
        self.export_type = export_type          # 내보낼 파일 타입 (txt, print, xlsx, email)
        self.path = path                        # 내보낼 경로 (default : desktop)
        # 내보낼 이름 (default : Vtubernews_날짜_시간)
        self.filename = filename + str(datetime.today().strftime('%m_%d_%H_%M'))
        self.mail_info = {'sender' : mail_addr,
                          'password' : "",
                          'rcver' : mail_addr,
                          'subject' : self.filename,
                        }

        # self.days = days                        # 며칠 전까지 수집할 것인가 (현재는 사용하지 않음)

        # parameters ======================================================================================
        self.pagenum = 0                                                        # 개념글 페이지 iteration
        self.today = datetime.today()                                           # 오늘 날짜
        self.user_agent = {'User-agent': 'Mozilla/5.0'}                         # 접속설정
        self.urlbase = "https://gall.dcinside.com/mgallery/board/lists/?id="    # 접속할 개념글 주소 베이스
        self.urlbase_view = "https://gall.dcinside.com/m/"                      # 내보낼 개념글 주소 베이스

        # performance test =================================================================================
        self.time_start = None

    ### 다음 개념글 페이지의 주소를 리턴한다.
    def get_next_page(self) :
        # 조합 방식 : https://gall.dcinside.com/mgallery/board/lists/?id= + kizunaai + &page= + 1 + &exception_mode=recommend
        self.pagenum += 1
        next_page_url = self.urlbase + self.gcode + "&page=" + str(self.pagenum) + "&exception_mode=recommend"
        return next_page_url

    ### URL에서 html을 받아온다.
    def get_html(self, url):
        _html = ""
        suc = False
        
        # 성공할 때까지 반복하며 html을 가져옴.
        while(suc == False):
            
            try:
                response = rq.get(url,headers=self.user_agent)
            
            except rq.exceptions.RequestException as e:
                time.sleep(3)
                continue

            if response.status_code == 200:
                suc = True
                _html = response.text

            else:
                suc = True
                _html = "<tbody><td>잘못된 주소 입니다.</td></tbody>"

        return _html

    ### 오늘의 날짜를 04.09 의 string으로 return.
    def get_today(self) :
        return str(self.today.strftime('%m.%d'))

    ### n일 전의 날짜를 04.09 의 string 형태로 return.
    def get_beforeday(self, nday) :
        return str((datetime.today()-timedelta(days=nday)).strftime('%m.%d'))

    ### html에서 얻어온 데이터를 파싱해 집어넣는다.
    def parse_data(self, html, newposts) :
        
        # 해당 페이지의 정보를 html로 읽어서
        soup = bs(html, 'html.parser')

        # 해당 페이지에 있는 모든 글의 정보를 리스트에 저장
        rawposts = soup.find("tbody").find_all("tr", class_="ub-content us-post")

        # 모든 글들의 리스트에서 필요한 정보만 추출.
        for rawpost in rawposts :

            # 만일 공지나 이벤트인 경우는 넘김.
            if (rawpost.find("td", class_="gall_subject").text == "이벤트") or (rawpost.find("td", class_="gall_subject").text == "공지") or (rawpost.find("td", class_="gall_subject").text == "설문") :
                continue

            # 보이스가 있는 경우는 넘김
            if rawpost.find("td", class_="gall_tit ub-word voice_tit") :
                continue

            # 해당한 키워드가 있는 경우에만
            if self.has_keyword(rawpost) :

                # newpost = [4.9 , 6809, 74, 냠냠, ㅇㅇ, 링크]
                newpost = []

                newpost.append(rawpost.find("td", class_="gall_date").text) # 날짜
                newpost.append(rawpost.find("td", class_="gall_count").text) # 조회수
                newpost.append(rawpost.find("td", class_="gall_recommend").text) # 추천수
                newpost.append(rawpost.find("td", class_="gall_tit ub-word").text[:-5][1:]) # 제목
                newpost.append(self.urlbase_view + self.gcode + "/" + rawpost.find("td", class_="gall_num").text) # 링크 

                # 글 모음에 집어넣음.
                newposts.append(newpost)

            # 키워드가 없는 경우는 넘김
            else :
                continue

            # 만일 최대 개수까지 채웠으면 함수 종료
            if self.is_full(newposts) :
                return

    ### 글의 제목에 지정된 키워드가 있는지를 판정한다.
    def has_keyword(self, title) :
        for word in self.list_keywords :
            if word in title.find("td", class_="gall_tit ub-word").text[:-5][1:] :
                return True
        return False

    ### 파싱한 정보가 최대 개수를 넘었는지 판정한다.
    def is_full(self, newposts) :
        return (len(newposts) >= int(self.maxpost))

    ### 저장할 파일 경로를 반환한다. (디폴트 바탕화면)
    def get_filepath(self, extension) :
        
        # 넘어온 값이 있으면
        if self.path : 
            return self.path + self.filename + extension
        
        # 넘어온 값이 없으면
        else : 
            filename = "\\" + self.filename + extension
            filepath = os.path.join(os.path.expanduser('~'),'Desktop') + filename 
            return filepath

    ### 파싱한 내용을 txt file로 export. 
    def export_txt(self, newposts) :
        
        with open(self.get_filepath(".txt"), "w") as f:
            f.write(newposts) # 링크    
            f.close()

    ### 파싱한 내용을 xlsx file로 export. 
    def export_xlsx(self, newposts) :

        write_wb = Workbook()
        write_ws = write_wb.create_sheet('시트1')
        write_ws = write_wb.active

        # 첫줄은 스키마
        write_ws.cell(row=1, column=1).value = '날짜'
        write_ws.cell(row=1, column=2).value = '조회수'
        write_ws.cell(row=1, column=3).value = '추천수'
        write_ws.cell(row=1, column=4).value = '제목'
        write_ws.cell(row=1, column=5).value = '링크'


        # 내용 입력    
        for i in range(len(newposts)) :
            for j in range(len(newposts[i])) :
                write_ws.cell(row=i+2, column=j+1).value = newposts[i][j]
                if j==4 :
                    write_ws.cell(row=i+2, column=j+1).style = "Hyperlink"

        # 너비조정
        # write_ws.column_dimensions['제목'].width = 55
        # write_ws.column_dimensions['링크'].width = 55

        # 저장
        write_wb.save(self.get_filepath(".xlsx"))

    # 파싱한 내용을 mail로 export.
    def export_mail(self, newposts) :
        
        # pw를 채워넣음
        self.mail_info['password'] = input('PW : ')

        # 메일 구성 (MIMEText)
        msg = MIMEMultipart("alternative")
        msg.set_charset('utf-8')

        # 정보 불러오기
        msg['Subject'] = self.mail_info['subject']
        msg['From'] = self.mail_info['sender']
        msg['To'] = self.mail_info['rcver']

        # html쓰기
        bodyPart = MIMEText(self.trans_html(newposts), 'html', 'utf-8')
        msg.attach( bodyPart )

        # 메일 서버 연결
        s=smtplib.SMTP( "smtp.gmail.com" , 587 )
        s.starttls() #TLS
        s.login( self.mail_info['sender'] , self.mail_info['password'] )
        s.sendmail( self.mail_info['sender'], self.mail_info['rcver'], msg.as_string() )
        s.close()

        print('Successfully Sent!')

    # 파싱한 내용을 html로 변환한다.
    def trans_html(self, newposts) :

        head = '''<head>

  <meta http-equiv="Content-Type" content="text/html; charset=UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <meta name="x-apple-disable-message-reformatting">
  <!--[if !mso]><!--><meta http-equiv="X-UA-Compatible" content="IE=edge"><!--<![endif]-->
  <title></title>
  
    <style type="text/css">
      @media only screen and (min-width: 620px) {
  .u-row {
    width: 600px !important;
  }
  .u-row .u-col {
    vertical-align: top;
  }

  .u-row .u-col-9p5 {
    width: 57px !important;
  }

  .u-row .u-col-9p74 {
    width: 58.44px !important;
  }

  .u-row .u-col-9p83 {
    width: 58.98px !important;
  }

  .u-row .u-col-10p01 {
    width: 60.06px !important;
  }

  .u-row .u-col-10p22 {
    width: 61.32px !important;
  }

  .u-row .u-col-10p33 {
    width: 61.98px !important;
  }

  .u-row .u-col-10p34 {
    width: 62.04px !important;
  }

  .u-row .u-col-10p38 {
    width: 62.28000000000001px !important;
  }

  .u-row .u-col-10p72 {
    width: 64.32px !important;
  }

  .u-row .u-col-10p88 {
    width: 65.28000000000002px !important;
  }

  .u-row .u-col-11 {
    width: 66px !important;
  }

  .u-row .u-col-58p82 {
    width: 352.92px !important;
  }

  .u-row .u-col-58p83 {
    width: 352.98px !important;
  }

  .u-row .u-col-59p18 {
    width: 355.08px !important;
  }

  .u-row .u-col-100 {
    width: 600px !important;
  }

}

@media (max-width: 620px) {
  .u-row-container {
    max-width: 100% !important;
    padding-left: 0px !important;
    padding-right: 0px !important;
  }
  .u-row .u-col {
    min-width: 320px !important;
    max-width: 100% !important;
    display: block !important;
  }
  .u-row {
    width: calc(100% - 40px) !important;
  }
  .u-col {
    width: 100% !important;
  }
  .u-col > div {
    margin: 0 auto;
  }
}
body {
  margin: 0;
  padding: 0;
}

table,
tr,
td {
  vertical-align: top;
  border-collapse: collapse;
}

p {
  margin: 0;
}

.ie-container table,
.mso-container table {
  table-layout: fixed;
}

* {
  line-height: inherit;
}

a[x-apple-data-detectors='true'] {
  color: inherit !important;
  text-decoration: none !important;
}

table, td { color: #000000; } a { color: #0000ee; text-decoration: underline; } @media (max-width: 480px) { #u_content_heading_17 .v-container-padding-padding { padding: 30px 10px 5px 20px !important; } #u_content_text_5 .v-container-padding-padding { padding: 10px 10px 10px 20px !important; } #u_content_text_6 .v-container-padding-padding { padding: 10px 10px 15px 20px !important; } #u_content_text_7 .v-container-padding-padding { padding: 10px 10px 30px 20px !important; } #u_content_heading_4 .v-container-padding-padding { padding: 30px 10px 10px !important; } #u_content_heading_6 .v-container-padding-padding { padding: 20px 10px 30px !important; } #u_content_heading_24 .v-container-padding-padding { padding: 20px 10px 30px !important; } #u_content_heading_25 .v-container-padding-padding { padding: 20px 10px 30px !important; } #u_content_heading_18 .v-container-padding-padding { padding: 30px 10px 10px !important; } #u_content_heading_20 .v-container-padding-padding { padding: 20px 10px 30px !important; } #u_content_heading_11 .v-container-padding-padding { padding: 30px 10px 10px !important; } #u_content_heading_13 .v-container-padding-padding { padding: 20px 10px 30px !important; } #u_content_heading_34 .v-container-padding-padding { padding: 30px 10px 10px !important; } #u_content_heading_36 .v-container-padding-padding { padding: 20px 10px 30px !important; } #u_content_heading_78 .v-container-padding-padding { padding: 30px 10px 10px !important; } #u_content_heading_80 .v-container-padding-padding { padding: 20px 10px 30px !important; } #u_content_heading_58 .v-container-padding-padding { padding: 30px 10px 10px !important; } #u_content_heading_60 .v-container-padding-padding { padding: 20px 10px 30px !important; } #u_content_heading_74 .v-container-padding-padding { padding: 30px 10px 10px !important; } #u_content_heading_76 .v-container-padding-padding { padding: 20px 10px 30px !important; } #u_content_heading_54 .v-container-padding-padding { padding: 30px 10px 10px !important; } #u_content_heading_56 .v-container-padding-padding { padding: 20px 10px 30px !important; } #u_content_heading_70 .v-container-padding-padding { padding: 30px 10px 10px !important; } #u_content_heading_72 .v-container-padding-padding { padding: 20px 10px 30px !important; } #u_content_heading_50 .v-container-padding-padding { padding: 30px 10px 10px !important; } #u_content_heading_52 .v-container-padding-padding { padding: 20px 10px 30px !important; } #u_content_heading_66 .v-container-padding-padding { padding: 30px 10px 10px !important; } #u_content_heading_68 .v-container-padding-padding { padding: 20px 10px 30px !important; } #u_content_heading_46 .v-container-padding-padding { padding: 30px 10px 10px !important; } #u_content_heading_48 .v-container-padding-padding { padding: 20px 10px 30px !important; } #u_content_heading_62 .v-container-padding-padding { padding: 30px 10px 10px !important; } #u_content_heading_64 .v-container-padding-padding { padding: 20px 10px 30px !important; } #u_content_heading_42 .v-container-padding-padding { padding: 30px 10px 10px !important; } #u_content_heading_44 .v-container-padding-padding { padding: 20px 10px 30px !important; } #u_content_heading_38 .v-container-padding-padding { padding: 30px 10px 10px !important; } #u_content_heading_40 .v-container-padding-padding { padding: 20px 10px 30px !important; } #u_content_heading_2 .v-container-padding-padding { padding: 30px 10px 5px 20px !important; } #u_content_text_4 .v-container-padding-padding { padding: 10px 10px 10px 20px !important; } #u_content_button_1 .v-size-width { width: auto !important; } #u_content_heading_21 .v-container-padding-padding { padding: 50px 10px 30px !important; } }
    </style>
  
  

<!--[if !mso]><!--><link href="https://fonts.googleapis.com/css?family=Montserrat:400,700&display=swap" rel="stylesheet" type="text/css"><link href="https://fonts.googleapis.com/css?family=Raleway:400,700&display=swap" rel="stylesheet" type="text/css"><link href="https://fonts.googleapis.com/css?family=Rubik:400,700&display=swap" rel="stylesheet" type="text/css"><!--<![endif]-->

</head>

<body class="clean-body u_body" style="margin: 0;padding: 0;-webkit-text-size-adjust: 100%;background-color: #ffffff;color: #000000">
  <!--[if IE]><div class="ie-container"><![endif]-->
  <!--[if mso]><div class="mso-container"><![endif]-->
  <table style="border-collapse: collapse;table-layout: fixed;border-spacing: 0;mso-table-lspace: 0pt;mso-table-rspace: 0pt;vertical-align: top;min-width: 320px;Margin: 0 auto;background-color: #ffffff;width:100%" cellpadding="0" cellspacing="0">
  <tbody>
  <tr style="vertical-align: top">
    <td style="word-break: break-word;border-collapse: collapse !important;vertical-align: top">
    <!--[if (mso)|(IE)]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td align="center" style="background-color: #ffffff;"><![endif]-->
    

<div class="u-row-container" style="padding: 0px;background-color: #26264f">
  <div class="u-row" style="Margin: 0 auto;min-width: 320px;max-width: 600px;overflow-wrap: break-word;word-wrap: break-word;word-break: break-word;background-color: transparent;">
    <div style="border-collapse: collapse;display: table;width: 100%;background-color: transparent;">
      <!--[if (mso)|(IE)]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding: 0px;background-color: #26264f;" align="center"><table cellpadding="0" cellspacing="0" border="0" style="width:600px;"><tr style="background-color: transparent;"><![endif]-->
      
<!--[if (mso)|(IE)]><td align="center" width="600" style="width: 600px;padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;" valign="top"><![endif]-->
<div class="u-col u-col-100" style="max-width: 320px;min-width: 600px;display: table-cell;vertical-align: top;">
  <div style="width: 100% !important;">
  <!--[if (!mso)&(!IE)]><!--><div style="padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;"><!--<![endif]-->
  
<table style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:10px;font-family:arial,helvetica,sans-serif;" align="left">
        
  <h1 style="margin: 0px; color: #ffffff; line-height: 140%; text-align: center; word-wrap: break-word; font-weight: normal; font-family: 'Raleway',sans-serif; font-size: 50px;">
    <strong>HoloNews</strong>
  </h1>

      </td>
    </tr>
  </tbody>
</table>

  <!--[if (!mso)&(!IE)]><!--></div><!--<![endif]-->
  </div>
</div>
<!--[if (mso)|(IE)]></td><![endif]-->
      <!--[if (mso)|(IE)]></tr></table></td></tr></table><![endif]-->
    </div>
  </div>
</div>



<div class="u-row-container" style="padding: 0px;background-color: transparent">
  <div class="u-row" style="Margin: 0 auto;min-width: 320px;max-width: 600px;overflow-wrap: break-word;word-wrap: break-word;word-break: break-word;background-color: #ffffff;">
    <div style="border-collapse: collapse;display: table;width: 100%;background-color: transparent;">
      <!--[if (mso)|(IE)]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding: 0px;background-color: transparent;" align="center"><table cellpadding="0" cellspacing="0" border="0" style="width:600px;"><tr style="background-color: #ffffff;"><![endif]-->
      
<!--[if (mso)|(IE)]><td align="center" width="600" style="width: 600px;padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;" valign="top"><![endif]-->
<div class="u-col u-col-100" style="max-width: 320px;min-width: 600px;display: table-cell;vertical-align: top;">
  <div style="width: 100% !important;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;">
  <!--[if (!mso)&(!IE)]><!--><div style="padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;"><!--<![endif]-->
  
<table style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:10px;font-family:arial,helvetica,sans-serif;" align="left">
        
<table width="100%" cellpadding="0" cellspacing="0" border="0">
  <tr>
    <td style="padding-right: 0px;padding-left: 0px;" align="center">
      
      <img align="center" border="0" src="images/image-1.jpeg" alt="Hero Image" title="Hero Image" style="outline: none;text-decoration: none;-ms-interpolation-mode: bicubic;clear: both;display: inline-block !important;border: none;height: auto;float: none;width: 100%;max-width: 500px;" width="500"/>
      
    </td>
  </tr>
</table>

      </td>
    </tr>
  </tbody>
</table>

  <!--[if (!mso)&(!IE)]><!--></div><!--<![endif]-->
  </div>
</div>
<!--[if (mso)|(IE)]></td><![endif]-->
      <!--[if (mso)|(IE)]></tr></table></td></tr></table><![endif]-->
    </div>
  </div>
</div>
<div class="u-row-container" style="padding: 0px;background-color: transparent">
  <div class="u-row" style="Margin: 0 auto;min-width: 320px;max-width: 600px;overflow-wrap: break-word;word-wrap: break-word;word-break: break-word;background-color: #6a71a8;">
    <div style="border-collapse: collapse;display: table;width: 100%;background-color: transparent;">
      <!--[if (mso)|(IE)]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding: 0px;background-color: transparent;" align="center"><table cellpadding="0" cellspacing="0" border="0" style="width:600px;"><tr style="background-color: #6a71a8;"><![endif]-->
      
<!--[if (mso)|(IE)]><td align="center" width="60" style="width: 60px;padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;" valign="top"><![endif]-->
<div class="u-col u-col-10p01" style="max-width: 320px;min-width: 60px;display: table-cell;vertical-align: top;">
  <div style="width: 100% !important;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;">
  <!--[if (!mso)&(!IE)]><!--><div style="padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;"><!--<![endif]-->
  
<table id="u_content_heading_4" style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:20px 10px 10px;font-family:arial,helvetica,sans-serif;" align="left">
        
  <h4 style="margin: 0px; color: #ffffff; line-height: 140%; text-align: center; word-wrap: break-word; font-weight: normal; font-family: 'Rubik',sans-serif; font-size: 16px;">
    <strong>Date</strong>
  </h4>

      </td>
    </tr>
  </tbody>
</table>

  <!--[if (!mso)&(!IE)]><!--></div><!--<![endif]-->
  </div>
</div>
<!--[if (mso)|(IE)]></td><![endif]-->
<!--[if (mso)|(IE)]><td align="center" width="353" style="width: 353px;padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;" valign="top"><![endif]-->
<div class="u-col u-col-58p82" style="max-width: 320px;min-width: 353px;display: table-cell;vertical-align: top;">
  <div style="width: 100% !important;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;">
  <!--[if (!mso)&(!IE)]><!--><div style="padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;"><!--<![endif]-->
  
<table style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:20px 10px 10px;font-family:arial,helvetica,sans-serif;" align="left">
        
  <h4 style="margin: 0px; color: #ffffff; line-height: 140%; text-align: center; word-wrap: break-word; font-weight: normal; font-family: 'Rubik',sans-serif; font-size: 16px;">
    <strong>Title</strong>
  </h4>

      </td>
    </tr>
  </tbody>
</table>

  <!--[if (!mso)&(!IE)]><!--></div><!--<![endif]-->
  </div>
</div>
<!--[if (mso)|(IE)]></td><![endif]-->
<!--[if (mso)|(IE)]><td align="center" width="62" style="width: 62px;padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;" valign="top"><![endif]-->
<div class="u-col u-col-10p34" style="max-width: 320px;min-width: 62px;display: table-cell;vertical-align: top;">
  <div style="width: 100% !important;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;">
  <!--[if (!mso)&(!IE)]><!--><div style="padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;"><!--<![endif]-->
  
<table id="u_content_heading_6" style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:20px 10px 15px;font-family:arial,helvetica,sans-serif;" align="left">
        
  <h4 style="margin: 0px; color: #ffffff; line-height: 140%; text-align: center; word-wrap: break-word; font-weight: normal; font-family: 'Rubik',sans-serif; font-size: 16px;">
    <strong>Rec</strong>
  </h4>

      </td>
    </tr>
  </tbody>
</table>

  <!--[if (!mso)&(!IE)]><!--></div><!--<![endif]-->
  </div>
</div>
<!--[if (mso)|(IE)]></td><![endif]-->
<!--[if (mso)|(IE)]><td align="center" width="66" style="width: 66px;padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;" valign="top"><![endif]-->
<div class="u-col u-col-11" style="max-width: 320px;min-width: 66px;display: table-cell;vertical-align: top;">
  <div style="width: 100% !important;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;">
  <!--[if (!mso)&(!IE)]><!--><div style="padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;"><!--<![endif]-->
  
<table id="u_content_heading_24" style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:20px 10px 15px;font-family:arial,helvetica,sans-serif;" align="left">
        
  <h4 style="margin: 0px; color: #ffffff; line-height: 140%; text-align: center; word-wrap: break-word; font-weight: normal; font-family: 'Rubik',sans-serif; font-size: 16px;">
    <strong>Click</strong>
  </h4>

      </td>
    </tr>
  </tbody>
</table>

  <!--[if (!mso)&(!IE)]><!--></div><!--<![endif]-->
  </div>
</div>
<!--[if (mso)|(IE)]></td><![endif]-->
<!--[if (mso)|(IE)]><td align="center" width="59" style="width: 59px;padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;" valign="top"><![endif]-->
<div class="u-col u-col-9p83" style="max-width: 320px;min-width: 59px;display: table-cell;vertical-align: top;">
  <div style="width: 100% !important;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;">
  <!--[if (!mso)&(!IE)]><!--><div style="padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;"><!--<![endif]-->
  
<table id="u_content_heading_25" style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:20px 10px 15px;font-family:arial,helvetica,sans-serif;" align="left">
        
  <h4 style="margin: 0px; color: #ffffff; line-height: 140%; text-align: center; word-wrap: break-word; font-weight: normal; font-family: 'Rubik',sans-serif; font-size: 16px;">
    <strong>Link</strong>
  </h4>

      </td>
    </tr>
  </tbody>
</table>

  <!--[if (!mso)&(!IE)]><!--></div><!--<![endif]-->
  </div>
</div>
<!--[if (mso)|(IE)]></td><![endif]-->
      <!--[if (mso)|(IE)]></tr></table></td></tr></table><![endif]-->
    </div>
  </div>
</div>
    '''
        tail = '''<div class="u-row-container" style="padding: 0px;background-color: transparent">
  <div class="u-row" style="Margin: 0 auto;min-width: 320px;max-width: 600px;overflow-wrap: break-word;word-wrap: break-word;word-break: break-word;background-color: #f0f5fa;">
    <div style="border-collapse: collapse;display: table;width: 100%;background-color: transparent;">
      <!--[if (mso)|(IE)]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding: 0px;background-color: transparent;" align="center"><table cellpadding="0" cellspacing="0" border="0" style="width:600px;"><tr style="background-color: #f0f5fa;"><![endif]-->
      
<!--[if (mso)|(IE)]><td align="center" width="600" style="width: 600px;padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;" valign="top"><![endif]-->
<div class="u-col u-col-100" style="max-width: 320px;min-width: 600px;display: table-cell;vertical-align: top;">
  <div style="width: 100% !important;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;">
  <!--[if (!mso)&(!IE)]><!--><div style="padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;"><!--<![endif]-->
  
<table id="u_content_heading_2" style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:40px 10px 0px 40px;font-family:arial,helvetica,sans-serif;" align="left">
        
  <h3 style="margin: 0px; color: #26264f; line-height: 140%; text-align: left; word-wrap: break-word; font-weight: normal; font-family: 'Montserrat',sans-serif; font-size: 26px;">
    <strong>That's the end of today!</strong>
  </h3>

      </td>
    </tr>
  </tbody>
</table>

<table id="u_content_text_4" style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:3px 40px 30px;font-family:arial,helvetica,sans-serif;" align="left">
        
  <div style="color: #7a7a7e; line-height: 170%; text-align: left; word-wrap: break-word;">
    <ul style="list-style-type: square;">
<li style="font-size: 14px; line-height: 23.8px;"><span style="font-size: 16px; line-height: 27.2px; font-family: Rubik, sans-serif;">Thanks for watching!</span></li>
</ul>
  </div>

      </td>
    </tr>
  </tbody>
</table>

<table id="u_content_button_1" style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:10px 10px 70px;font-family:arial,helvetica,sans-serif;" align="left">
        
<div align="center">
  <!--[if mso]><table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-spacing: 0; border-collapse: collapse; mso-table-lspace:0pt; mso-table-rspace:0pt;font-family:arial,helvetica,sans-serif;"><tr><td style="font-family:arial,helvetica,sans-serif;" align="center"><v:roundrect xmlns:v="urn:schemas-microsoft-com:vml" xmlns:w="urn:schemas-microsoft-com:office:word" href="https://gall.dcinside.com/mgallery/board/lists?id=kizunaai&exception_mode=recommend" style="height:64px; v-text-anchor:middle; width:435px;" arcsize="1.5%" stroke="f" fillcolor="#6a71a8"><w:anchorlock/><center style="color:#FFFFFF;font-family:arial,helvetica,sans-serif;"><![endif]-->
    <a href="https://gall.dcinside.com/mgallery/board/lists?id=kizunaai&exception_mode=recommend" target="_blank" class="v-size-width" style="box-sizing: border-box;display: inline-block;font-family:arial,helvetica,sans-serif;text-decoration: none;-webkit-text-size-adjust: none;text-align: center;color: #FFFFFF; background-color: #6a71a8; border-radius: 1px;-webkit-border-radius: 1px; -moz-border-radius: 1px; width:75%; max-width:100%; overflow-wrap: break-word; word-break: break-word; word-wrap:break-word; mso-border-alt: none;">
      <span style="display:block;padding:21px 20px;line-height:120%;"><span style="font-size: 18px; line-height: 21.6px; font-family: Rubik, sans-serif;"><span style="line-height: 21.6px; font-size: 18px;">Click Below to Access Gallery</span></span></span>
    </a>
  <!--[if mso]></center></v:roundrect></td></tr></table><![endif]-->
</div>

      </td>
    </tr>
  </tbody>
</table>

  <!--[if (!mso)&(!IE)]><!--></div><!--<![endif]-->
  </div>
</div>
<!--[if (mso)|(IE)]></td><![endif]-->
      <!--[if (mso)|(IE)]></tr></table></td></tr></table><![endif]-->
    </div>
  </div>
</div>



<div class="u-row-container" style="padding: 0px;background-color: #26264f">
  <div class="u-row" style="Margin: 0 auto;min-width: 320px;max-width: 600px;overflow-wrap: break-word;word-wrap: break-word;word-break: break-word;background-color: transparent;">
    <div style="border-collapse: collapse;display: table;width: 100%;background-color: transparent;">
      <!--[if (mso)|(IE)]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding: 0px;background-color: #26264f;" align="center"><table cellpadding="0" cellspacing="0" border="0" style="width:600px;"><tr style="background-color: transparent;"><![endif]-->
      
<!--[if (mso)|(IE)]><td align="center" width="600" style="width: 600px;padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;" valign="top"><![endif]-->
<div class="u-col u-col-100" style="max-width: 320px;min-width: 600px;display: table-cell;vertical-align: top;">
  <div style="width: 100% !important;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;">
  <!--[if (!mso)&(!IE)]><!--><div style="padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;"><!--<![endif]-->
  
<table id="u_content_heading_21" style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:50px 10px 20px;font-family:arial,helvetica,sans-serif;" align="left">
        
  <h1 style="margin: 0px; color: #ffffff; line-height: 140%; text-align: center; word-wrap: break-word; font-weight: normal; font-family: 'Montserrat',sans-serif; font-size: 31px;">
    Thank You for being with Us!
  </h1>

      </td>
    </tr>
  </tbody>
</table>

<table style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:20px 10px 50px;font-family:arial,helvetica,sans-serif;" align="left">
        
  <div style="color: #d4d4d4; line-height: 180%; text-align: center; word-wrap: break-word;">
    <p style="font-size: 14px; line-height: 180%;"><span style="font-family: Rubik, sans-serif; font-size: 14px; line-height: 25.2px;">If you have any questions, feel free message us at support@mailus.com. </span><br /><span style="font-family: Rubik, sans-serif; font-size: 14px; line-height: 25.2px;">All rights reserved. Update email preferences or unsubscribe.</span><br /><span style="font-family: Rubik, sans-serif; font-size: 14px; line-height: 25.2px;">123-456-7890</span><br /><span style="font-family: Rubik, sans-serif; font-size: 14px; line-height: 25.2px;">San Francisco, CA. United States</span><br /><span style="font-family: Rubik, sans-serif; font-size: 14px; line-height: 25.2px;">Terms of use | Privacy Policy&nbsp;</span></p>
  </div>

      </td>
    </tr>
  </tbody>
</table>

  <!--[if (!mso)&(!IE)]><!--></div><!--<![endif]-->
  </div>
</div>
<!--[if (mso)|(IE)]></td><![endif]-->
      <!--[if (mso)|(IE)]></tr></table></td></tr></table><![endif]-->
    </div>
  </div>
</div>



<div class="u-row-container" style="padding: 0px;background-color: transparent">
  <div class="u-row" style="Margin: 0 auto;min-width: 320px;max-width: 600px;overflow-wrap: break-word;word-wrap: break-word;word-break: break-word;background-color: transparent;">
    <div style="border-collapse: collapse;display: table;width: 100%;background-color: transparent;">
      <!--[if (mso)|(IE)]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding: 0px;background-color: transparent;" align="center"><table cellpadding="0" cellspacing="0" border="0" style="width:600px;"><tr style="background-color: transparent;"><![endif]-->
      
<!--[if (mso)|(IE)]><td align="center" width="600" style="width: 600px;padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;" valign="top"><![endif]-->
<div class="u-col u-col-100" style="max-width: 320px;min-width: 600px;display: table-cell;vertical-align: top;">
  <div style="width: 100% !important;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;">
  <!--[if (!mso)&(!IE)]><!--><div style="padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;"><!--<![endif]-->
  
<table style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:20px 10px;font-family:arial,helvetica,sans-serif;" align="left">
        
  <div style="color: #95a5a6; line-height: 140%; text-align: center; word-wrap: break-word;">
    <p style="font-size: 14px; line-height: 140%;"><span style="font-family: Rubik, sans-serif; font-size: 14px; line-height: 19.6px;">&copy; 20XX Company. All Rights Reserved.</span></p>
  </div>

      </td>
    </tr>
  </tbody>
</table>

  <!--[if (!mso)&(!IE)]><!--></div><!--<![endif]-->
  </div>
</div>
<!--[if (mso)|(IE)]></td><![endif]-->
      <!--[if (mso)|(IE)]></tr></table></td></tr></table><![endif]-->
    </div>
  </div>
</div>


    <!--[if (mso)|(IE)]></td></tr></table><![endif]-->
    </td>
  </tr>
  </tbody>
  </table>
  <!--[if mso]></div><![endif]-->
  <!--[if IE]></div><![endif]-->
</body>

</html>
    

    
    '''
        mid1 = '''<div class="u-row-container" style="padding: 0px;background-color: transparent">
  <div class="u-row" style="Margin: 0 auto;min-width: 320px;max-width: 600px;overflow-wrap: break-word;word-wrap: break-word;word-break: break-word;background-color: #f4f8fb;">
    <div style="border-collapse: collapse;display: table;width: 100%;background-color: transparent;">
      <!--[if (mso)|(IE)]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding: 0px;background-color: transparent;" align="center"><table cellpadding="0" cellspacing="0" border="0" style="width:600px;"><tr style="background-color: #f4f8fb;"><![endif]-->
      
<!--[if (mso)|(IE)]><td align="center" width="58" style="width: 58px;padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;" valign="top"><![endif]-->
<div class="u-col u-col-9p74" style="max-width: 320px;min-width: 58px;display: table-cell;vertical-align: top;">
  <div style="width: 100% !important;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;">
  <!--[if (!mso)&(!IE)]><!--><div style="padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;"><!--<![endif]-->
  
<table id="u_content_heading_18" style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:20px 10px 15px;font-family:arial,helvetica,sans-serif;" align="left">
        
  <h4 style="margin: 0px; color: #6a71a8; line-height: 140%; text-align: center; word-wrap: break-word; font-weight: normal; font-family: 'Rubik',sans-serif; font-size: 13px;">
    {date}
  </h4>

      </td>
    </tr>
  </tbody>
</table>

  <!--[if (!mso)&(!IE)]><!--></div><!--<![endif]-->
  </div>
</div>
<!--[if (mso)|(IE)]></td><![endif]-->
<!--[if (mso)|(IE)]><td align="center" width="353" style="width: 353px;padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;" valign="top"><![endif]-->
<div class="u-col u-col-58p83" style="max-width: 320px;min-width: 353px;display: table-cell;vertical-align: top;">
  <div style="width: 100% !important;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;">
  <!--[if (!mso)&(!IE)]><!--><div style="padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;"><!--<![endif]-->
  
<table style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:20px 10px 10px;font-family:arial,helvetica,sans-serif;" align="left">
        
  <h4 style="margin: 0px; color: #6a71a8; line-height: 150%; text-align: center; word-wrap: break-word; font-weight: normal; font-family: 'Rubik',sans-serif; font-size: 15px;">
    {title}
  </h4>

      </td>
    </tr>
  </tbody>
</table>

  <!--[if (!mso)&(!IE)]><!--></div><!--<![endif]-->
  </div>
</div>
<!--[if (mso)|(IE)]></td><![endif]-->
<!--[if (mso)|(IE)]><td align="center" width="62" style="width: 62px;padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;" valign="top"><![endif]-->
<div class="u-col u-col-10p33" style="max-width: 320px;min-width: 62px;display: table-cell;vertical-align: top;">
  <div style="width: 100% !important;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;">
  <!--[if (!mso)&(!IE)]><!--><div style="padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;"><!--<![endif]-->
  
<table id="u_content_heading_20" style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:20px 10px 30px;font-family:arial,helvetica,sans-serif;" align="left">
        
  <h4 style="margin: 0px; color: #6a71a8; line-height: 140%; text-align: center; word-wrap: break-word; font-weight: normal; font-family: 'Rubik',sans-serif; font-size: 14px;">
    {rec}
  </h4>

      </td>
    </tr>
  </tbody>
</table>

  <!--[if (!mso)&(!IE)]><!--></div><!--<![endif]-->
  </div>
</div>
<!--[if (mso)|(IE)]></td><![endif]-->
<!--[if (mso)|(IE)]><td align="center" width="65" style="width: 65px;padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;" valign="top"><![endif]-->
<div class="u-col u-col-10p88" style="max-width: 320px;min-width: 65px;display: table-cell;vertical-align: top;">
  <div style="width: 100% !important;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;">
  <!--[if (!mso)&(!IE)]><!--><div style="padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;"><!--<![endif]-->
  
<table style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:20px 10px 10px;font-family:arial,helvetica,sans-serif;" align="left">
        
  <h4 style="margin: 0px; color: #6a71a8; line-height: 140%; text-align: center; word-wrap: break-word; font-weight: normal; font-family: 'Rubik',sans-serif; font-size: 14px;">
    {click}
  </h4>

      </td>
    </tr>
  </tbody>
</table>

  <!--[if (!mso)&(!IE)]><!--></div><!--<![endif]-->
  </div>
</div>
<!--[if (mso)|(IE)]></td><![endif]-->
<!--[if (mso)|(IE)]><td align="center" width="61" style="width: 61px;padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;" valign="top"><![endif]-->
<div class="u-col u-col-10p22" style="max-width: 320px;min-width: 61px;display: table-cell;vertical-align: top;">
  <div style="width: 100% !important;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;">
  <!--[if (!mso)&(!IE)]><!--><div style="padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;"><!--<![endif]-->
  
<table style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:4px;font-family:arial,helvetica,sans-serif;" align="left">
        
<div align="center">
  <!--[if mso]><table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-spacing: 0; border-collapse: collapse; mso-table-lspace:0pt; mso-table-rspace:0pt;font-family:arial,helvetica,sans-serif;"><tr><td style="font-family:arial,helvetica,sans-serif;" align="center"><v:roundrect xmlns:v="urn:schemas-microsoft-com:vml" xmlns:w="urn:schemas-microsoft-com:office:word" href="https://gall.dcinside.com/m/kizunaai/5193521" style="height:42px; v-text-anchor:middle; width:52px;" arcsize="24%" stroke="f" fillcolor="#7686bd"><w:anchorlock/><center style="color:#FFFFFF;font-family:arial,helvetica,sans-serif;"><![endif]-->
    <a href={link} target="_blank" class="v-size-width" style="box-sizing: border-box;display: inline-block;font-family:arial,helvetica,sans-serif;text-decoration: none;-webkit-text-size-adjust: none;text-align: center;color: #FFFFFF; background-color: #7686bd; border-radius: 10px;-webkit-border-radius: 10px; -moz-border-radius: 10px; width:auto; max-width:100%; overflow-wrap: break-word; word-break: break-word; word-wrap:break-word; mso-border-alt: none;">
      <span style="display:block;padding:11px 20px;line-height:140%;"><span style="font-family: Rubik, sans-serif; font-size: 14px; line-height: 14px;"><strong><span style="font-size: 12px; line-height: 12px;">▶</span></strong></span></span>
    </a>
  <!--[if mso]></center></v:roundrect></td></tr></table><![endif]-->
</div>

      </td>
    </tr>
  </tbody>
</table>

  <!--[if (!mso)&(!IE)]><!--></div><!--<![endif]-->
  </div>
</div>
<!--[if (mso)|(IE)]></td><![endif]-->
      <!--[if (mso)|(IE)]></tr></table></td></tr></table><![endif]-->
    </div>
  </div>
</div>

    
    '''
        mid2 = '''<div class="u-row-container" style="padding: 0px;background-color: transparent">
  <div class="u-row" style="Margin: 0 auto;min-width: 320px;max-width: 600px;overflow-wrap: break-word;word-wrap: break-word;word-break: break-word;background-color: #ffffff;">
    <div style="border-collapse: collapse;display: table;width: 100%;background-color: transparent;">
      <!--[if (mso)|(IE)]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding: 0px;background-color: transparent;" align="center"><table cellpadding="0" cellspacing="0" border="0" style="width:600px;"><tr style="background-color: #ffffff;"><![endif]-->
      
<!--[if (mso)|(IE)]><td align="center" width="57" style="width: 57px;padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;" valign="top"><![endif]-->
<div class="u-col u-col-9p5" style="max-width: 320px;min-width: 57px;display: table-cell;vertical-align: top;">
  <div style="width: 100% !important;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;">
  <!--[if (!mso)&(!IE)]><!--><div style="padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;"><!--<![endif]-->
  
<table id="u_content_heading_11" style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:20px 10px 10px;font-family:arial,helvetica,sans-serif;" align="left">
        
  <h4 style="margin: 0px; color: #6a71a8; line-height: 140%; text-align: center; word-wrap: break-word; font-weight: normal; font-family: 'Rubik',sans-serif; font-size: 13px;">
    {date}
  </h4>

      </td>
    </tr>
  </tbody>
</table>

  <!--[if (!mso)&(!IE)]><!--></div><!--<![endif]-->
  </div>
</div>
<!--[if (mso)|(IE)]></td><![endif]-->
<!--[if (mso)|(IE)]><td align="center" width="355" style="width: 355px;padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;" valign="top"><![endif]-->
<div class="u-col u-col-59p18" style="max-width: 320px;min-width: 355px;display: table-cell;vertical-align: top;">
  <div style="width: 100% !important;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;">
  <!--[if (!mso)&(!IE)]><!--><div style="padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;"><!--<![endif]-->
  
<table style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:20px 10px 10px;font-family:arial,helvetica,sans-serif;" align="left">
        
  <h4 style="margin: 0px; color: #6a71a8; line-height: 140%; text-align: center; word-wrap: break-word; font-weight: normal; font-family: 'Rubik',sans-serif; font-size: 14px;">
    {title}
  </h4>

      </td>
    </tr>
  </tbody>
</table>

  <!--[if (!mso)&(!IE)]><!--></div><!--<![endif]-->
  </div>
</div>
<!--[if (mso)|(IE)]></td><![endif]-->
<!--[if (mso)|(IE)]><td align="center" width="61" style="width: 61px;padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;" valign="top"><![endif]-->
<div class="u-col u-col-10p22" style="max-width: 320px;min-width: 61px;display: table-cell;vertical-align: top;">
  <div style="width: 100% !important;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;">
  <!--[if (!mso)&(!IE)]><!--><div style="padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;"><!--<![endif]-->
  
<table id="u_content_heading_13" style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:20px 10px 30px;font-family:arial,helvetica,sans-serif;" align="left">
        
  <h4 style="margin: 0px; color: #6a71a8; line-height: 140%; text-align: center; word-wrap: break-word; font-weight: normal; font-family: 'Rubik',sans-serif; font-size: 14px;">
    {rec}
  </h4>

      </td>
    </tr>
  </tbody>
</table>

  <!--[if (!mso)&(!IE)]><!--></div><!--<![endif]-->
  </div>
</div>
<!--[if (mso)|(IE)]></td><![endif]-->
<!--[if (mso)|(IE)]><td align="center" width="64" style="width: 64px;padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;" valign="top"><![endif]-->
<div class="u-col u-col-10p72" style="max-width: 320px;min-width: 64px;display: table-cell;vertical-align: top;">
  <div style="width: 100% !important;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;">
  <!--[if (!mso)&(!IE)]><!--><div style="padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;"><!--<![endif]-->
  
<table style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:20px 10px 10px;font-family:arial,helvetica,sans-serif;" align="left">
        
  <h4 style="margin: 0px; color: #6a71a8; line-height: 140%; text-align: center; word-wrap: break-word; font-weight: normal; font-family: 'Rubik',sans-serif; font-size: 14px;">
    {click}
  </h4>

      </td>
    </tr>
  </tbody>
</table>

  <!--[if (!mso)&(!IE)]><!--></div><!--<![endif]-->
  </div>
</div>
<!--[if (mso)|(IE)]></td><![endif]-->
<!--[if (mso)|(IE)]><td align="center" width="62" style="width: 62px;padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;" valign="top"><![endif]-->
<div class="u-col u-col-10p38" style="max-width: 320px;min-width: 62px;display: table-cell;vertical-align: top;">
  <div style="width: 100% !important;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;">
  <!--[if (!mso)&(!IE)]><!--><div style="padding: 0px;border-top: 0px solid transparent;border-left: 0px solid transparent;border-right: 0px solid transparent;border-bottom: 0px solid transparent;border-radius: 0px;-webkit-border-radius: 0px; -moz-border-radius: 0px;"><!--<![endif]-->
  
<table style="font-family:arial,helvetica,sans-serif;" role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0">
  <tbody>
    <tr>
      <td class="v-container-padding-padding" style="overflow-wrap:break-word;word-break:break-word;padding:4px;font-family:arial,helvetica,sans-serif;" align="left">
        
<div align="center">
  <!--[if mso]><table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-spacing: 0; border-collapse: collapse; mso-table-lspace:0pt; mso-table-rspace:0pt;font-family:arial,helvetica,sans-serif;"><tr><td style="font-family:arial,helvetica,sans-serif;" align="center"><v:roundrect xmlns:v="urn:schemas-microsoft-com:vml" xmlns:w="urn:schemas-microsoft-com:office:word" href="https://gall.dcinside.com/m/kizunaai/5193521" style="height:42px; v-text-anchor:middle; width:52px;" arcsize="24%" stroke="f" fillcolor="#7686bd"><w:anchorlock/><center style="color:#FFFFFF;font-family:arial,helvetica,sans-serif;"><![endif]-->
    <a href={link} target="_blank" class="v-size-width" style="box-sizing: border-box;display: inline-block;font-family:arial,helvetica,sans-serif;text-decoration: none;-webkit-text-size-adjust: none;text-align: center;color: #FFFFFF; background-color: #7686bd; border-radius: 10px;-webkit-border-radius: 10px; -moz-border-radius: 10px; width:auto; max-width:100%; overflow-wrap: break-word; word-break: break-word; word-wrap:break-word; mso-border-alt: none;">
      <span style="display:block;padding:11px 20px;line-height:140%;"><span style="font-family: Rubik, sans-serif; font-size: 14px; line-height: 14px;"><strong><span style="font-size: 12px; line-height: 12px;">▶</span></strong></span></span>
    </a>
  <!--[if mso]></center></v:roundrect></td></tr></table><![endif]-->
</div>

      </td>
    </tr>
  </tbody>
</table>

  <!--[if (!mso)&(!IE)]><!--></div><!--<![endif]-->
  </div>
</div>
<!--[if (mso)|(IE)]></td><![endif]-->
      <!--[if (mso)|(IE)]></tr></table></td></tr></table><![endif]-->
    </div>
  </div>
</div>
    
    
    '''

        # 소식을 결합하여
        for post in newposts :
            if newposts.index(post)%2 == 0 :
                head += mid1.format(date=post[0], title=post[3], rec=post[2], click=post[1], link=post[4])
            else :
                head += mid2.format(date=post[0], title=post[3], rec=post[2], click=post[1], link=post[4])

        # 완성된 html로 리턴
        body = head + tail
        return body


    ### 파싱한 리스트를 콘솔창에 프린트(확인용)
    def print_list(self, newposts) :
        for post in newposts :
            print(post[0])
            print("조회수 : {} / 추천수 : {}".format(post[1], post[2]))
            print(post[3])
            print(post[4])
            print()

    ### 작업 처리시간을 반환한다.
    def get_perf_time(self) :
        return time.time() - self.time_start

    # 메인 함수.
    def start(self) :       

        # 0) 초기화
        newposts = []
        self.time_start = time.time()

        # 1) 글을 모으기
        while not self.is_full(newposts) :                      # 정해진 개수만큼 수집하기 전까지는
            newurl = self.get_next_page()                       # 다음 페이지의 주소를 구해
            self.parse_data(self.get_html(newurl), newposts)    # 받아온 정보를 newposts에 파싱
        
        # 2) 파일로 내보내기
        if self.export_type == 'xlsx' :     # 엑셀로
            self.export_xlsx(newposts)
        elif self.export_type == 'print' :  # 콘솔창으로
            self.print_list(newposts)
        elif self.export_type == 'mail' :   # 메일로
            self.export_mail(newposts)
        else :
            print('알 수 없는 타입입니다.')

        # 3) 퍼포먼스 측정
        print("소요 시간 : {} sec".format(round(self.get_perf_time(), 3)))


if __name__ == "__main__" :

    # ========================================================== config
    # 갤 코드를 입력하세요.
    gcode = 'kizunaai'    
    
    # 최대 몇 개의 post를 가져옵니까?
    maxpost = 30

    # xlsx, mail 중 선택
    export_type = "mail"

    # (엑셀일 경우) 파일을 내보낼 경로를 입력하세요.
    path = ""

    # (메일일 경우) 파일을 전송할 메일을 입력하세요.
    addr = "joonery79@gmail.com"
    
    # 검색할 키워드를 입력하세요.
    list_keywords = [
        # EN
        'EN', '2기','카운슬', 'myth',

        # 칼리
        '모리', '칼리', '데몬다이스', '사신',

        # 키아라                
        '키아라', '치킨', '불닭', '타카나시', '점장', 'KFP', '케키',

        # 아메                
        '아메', '왓슨', '슨상', '스몰', 'smol', 'ame',

        # 이나
        '타코', '이나', '희주', '무너', '문어',

        # 구라
        '상어', '구라',

        # 아이리스
        '아이리스', 'Rys', 'rys', '나미린', '희망',

        # 크로니
        '시계', '크로', '도토부',

        # 베이
        '땃쥐', '베이', '벨즈', '햄스터',

        # 무메이
        '나나시', '무메이', '우흥', '토리', '자폐',

        # 파우나
        '파우나', '자연', '리프', '마망', '비건', '릴리', '에디', '새플링',

        # 사나
        '사나', '흑인',
    ]


    gcode_genshin = 'onshinproject'
    kw_genshin = ['감우', '신학']

    # ========================================================== 실행
    alimi = KgallAlimi(gcode_genshin, kw_genshin, maxpost, export_type, mail_addr=addr)
    alimi.start()
