import requests as rq
import time
from bs4 import BeautifulSoup as bs


# # 되는주소
# response = rq.get("https://gall.dcinside.com/m/kizunaai/5190056")

# # 안되는 주소
# # response = rq.get('https://e-hentai.org/')

# print(response.status_code)
# print(response.text)

# bs1 = bs(response.text, 'html.parser')
# print(bs1, type(bs1))



# 200이 성공.
user_agent = {'User-agent': 'Mozilla/5.0'}
def get_html(url):
   _html = ""
   suc = False
   
   # 성공할 때까지 반복하며 html을 가져옴.
   while(suc == False):
       
       try:
        response = rq.get(url,headers=user_agent)
       
       except rq.exceptions.RequestException as e:
           time.sleep(3)
           continue

       if response.status_code == 200:
           suc = True
           _html = response.text

       else:
           suc = True
           _html = "<tbody><td>잘못된 주소 입니다.</td></tbody>"

   return _html


def ex2(html):

    soup = bs(html, 'html.parser')
    newposts = []

    # 유저가 쓴 글 목록의 정보를 모두 불러와 리스트로 저장.
    rawposts = soup.find("tbody").find_all("tr", class_="ub-content us-post")

    # 그 리스트에서 정보만 추출.
    for i in rawposts :
        newpost = []
        
        # newpost에 임시로 넣어
        # newpost = [4.9 , 6809, 74, 냠냠, ㅇㅇ, 링크]

        # newpost.append(i.find_all("td", class_="gall_date")) # 이렇게 하면 해당 행이 다 뽑힘.

        newpost.append(i.find("td", class_="gall_date").text) # 날짜/시간
        newpost.append(i.find("td", class_="gall_count").text) # 조회수
        newpost.append(i.find("td", class_="gall_recommend").text) # 추천수
        newpost.append(i.find("td", class_="gall_tit ub-word").text[:-6][1:]) # 제목
        newpost.append("https://gall.dcinside.com/m/kizunaai/" + i.find("td", class_="gall_num").text) # 링크 (조합해서 만들것)

        # 글 모음에 집어넣음.
        newposts.append(newpost) 


    # 출력.
    for i in newposts :
        print(i[0])
        print("조회수 : {} / 추천수 : {}".format(i[1], i[2]))
        print(i[3])
        print(i[4])
        print()
    
    print(len(newposts))


# url = "https://gall.dcinside.com/mgallery/board/lists/?id=kizunaai&page=1&exception_mode=recommend"
# print(ex2(get_html(url)))


# from datetime import datetime, timedelta

# # 오늘을 string으로 변환.
# print(str(datetime.today().strftime('%m.%d')))

# # 어제를 구하는 방법. days= n으로 하면 n일 전이 나온다. 이걸 str로 바꿔주면 돼지~
# print( (datetime.today() - timedelta(days=1)).strftime('%m.%d') )


# def is_full(newposts, maxpost) :
#     return (len(newposts) >= int(maxpost))

# a = []
# maxpost = 100
# print(is_full(a,maxpost))

import os
from openpyxl import Workbook

name = '냠'
filename = "\\" + name +  ".xlsx"      # 캬루야사랑해.txt란 이름으로
filepath = os.path.join(os.path.expanduser('~'),'Desktop') + filename 

listing = [[1,2,3],[1,2,3],[1,2,3]]

# with open(filepath, "w") as f:
#     for i in listing :
#         f.write("".format(i))
#         f.write('')
#     f.close()

write_wb = Workbook()
write_ws = write_wb.create_sheet('시트1')

write_ws = write_wb.active

for i in range(len(listing)) :
    for j in range(len(listing[i])) :
        write_ws.cell(row=i+1, column=j+1).value = listing[i][j]

write_wb.save(filepath)

# 이름이 있는 시트를 생성


# def f1() :
#     for i in range(10) :
#         print("i={}".format(i))
#         f2()

# def f2() :
#     for j in range(5) :
#         print("j={}".format(j))
#         if j > 2 :
#             break

# f1()





# 한 글의 html 정보 ============================================================
# </tr>
# <tr class="ub-content us-post" data-no="5181384" data-type="icon_recomtxt">
# <td class="gall_num">5181384</td>
# <td class="gall_subject">일반</td>
# <td class="gall_tit ub-word">
    # <a href="/mgallery/board/view/?id=kizunaai&amp;no=5181384&amp;exception_mode=recommend&amp;page=1" view-msg=""><em class="icon_img icon_recomtxt"></em>에펙가이지들 존나 신났네 그냥</a>
    # <a class="reply_numbox" href="https://gall.dcinside.com/mgallery/board/view/?id=kizunaai&amp;no=5181384&amp;t=cv&amp;exception_mode=recommend&amp;page=1"><span class="reply_num">[13]</span></a> </td>
# <td class="gall_writer ub-writer" data-ip="211.178" data-loc="list" data-nick="ㅇㅇ" data-uid="">
# <span class="nickname" title="ㅇㅇ"><em>ㅇㅇ</em></span><span class="ip">(211.178)</span> </td>
# <td class="gall_date" title="2022-04-09 03:45:02">04.09</td>
# <td class="gall_count">6809</td>
# <td class="gall_recommend">74</td>
# </tr>





# 추출한 html 정보 ============================================================
# [<td class="gall_date" title="2022-04-09 10:22:38">04.09</td>]
# [<td class="gall_count">9868</td>]
# [<td class="gall_recommend">59</td>]
# [<td class="gall_tit ub-word">
# <a href="/mgallery/board/view/?id=kizunaai&amp;no=5182678&amp;exception_mode=recommend&amp;page=1" view-msg=""><em class="icon_img icon_recomimg"></em>아ㅏㅏㅏㅏㅏㅏㅏㅏㅏㅏㅏㅏ</a>
# <a class="reply_numbox" href="https://gall.dcinside.com/mgallery/board/view/?id=kizunaai&amp;no=5182678&amp;t=cv&amp;exception_mode=recommend&amp;page=1"><span class="reply_num">[76]</span></a> </td>]
# [<td class="gall_num">5182678</td>]

# <td class="gall_subject">이벤트</td> : 특정 경우에는 제외하도록.

# <tr class="ub-content us-post" data-no="5179591" data-type="icon_recomimg">
# <td class="gall_num">5179591</td>
# <td class="gall_subject">이벤트</td>
# <td class="gall_tit ub-word voice_tit">
# <a href="/mgallery/board/view/?id=kizunaai&amp;no=5179591&amp;exception_mode=recommend&amp;page=2" view-msg=""><em class="icon_img icon_recomimg"></em>아ㅏㅏㅏㅏㅏㅏㅏㅏㅏㅏㅏㅏ</a><em class="sp_img icon_voice_tit"></em><a class="reply_numbox" href="https://gall.dcinside.com/mgallery/board/view/?id=kizunaai&amp;no=5179591&amp;t=cv&amp;exception_mode=recommend&amp;page=2"><span class="reply_num">[110/7]</span></a> </td>
# <td class="gall_writer ub-writer" data-ip="" data-loc="list" data-nick="Inokawa" data-uid="dqww001"><span class="nickname in" title="Inokawa"><em>Inokawa</em></span><a class="writer_nikcon"><img alt="갤로그로 이동합니다." border="0" height="11" onclick="window.open('//gallog.dcinside.com/dqww001');" src="https://nstatic.dcinside.com/dc/w/images/nik.gif" style="margin-left:2px;cursor:pointer;" title="dqww0** : 갤로그로 이동합니다." width="12"/></a> </td>
# <td class="gall_date" title="2022-04-09 00:51:51">04.09</td><td class="gall_count">5657</td><td class="gall_recommend">66</td></tr>